"""把单只全历史 CSV 出成带配色的 Excel —— 只做呈现,不改任何计算值。

用法
----
    OXQ_HIST_CODE=688347 uv run python rps_pool_study/stock_history_xlsx.py

输入 `{OUT}/stock_history_{code}.csv`(由 `template_20260828.py` 的
`OXQ_STOCK_HISTORY` 模式产出),输出同名 `.xlsx`。

配色规则(与 `pool_20260828_xlsx.py` 同一套色板,便于两边对照)
------------------------------------------------------------
- **整行底色**看「统一信号」:=1 的行整行浅红,让信号段在 700 行里一眼可见;
- 信号类型:强确认 红 / 标准确认 黄 / 观察级 蓝;
- 平台信号:平台突破(研究)橙 / 平台观察 浅黄;
- 触发状态:新触发 绿 / 持续 蓝;
- 周线五态、质量档沿用原色板;
- **每一个信号段的第一行加粗上边框**,直接把 13 个段切开。

另出一张「信号段」页:把统一信号=1 的连续段汇总成一行一段,
含起止日、天数、段内收盘涨跌幅 —— 这是 700 行里最该先看的东西。

**本文件不构成任何买入建议。**
"""

from __future__ import annotations

import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.environ.get("OXQ_OUT_DIR", "/home/user/oxq-panel")
CODE = os.environ.get("OXQ_HIST_CODE", "688347")
NAVY = PatternFill("solid", fgColor="17365D")
HF = Font(color="FFFFFF", bold=True, size=10)
ROWSIG = PatternFill("solid", fgColor="FDE9E9")      # 统一信号=1 的整行底色
FILL = {
    "信号类型": {"强确认": "FFC7CE", "标准确认": "FFEB9C", "观察级": "DDEBF7"},
    "平台信号": {"平台突破（研究）": "F8CBAD", "平台观察": "FFF2CC"},
    "案例展示分层_质量": {"高": "C6EFCE", "低": "FFC7CE", "缺失": "E7E6E6"},
    "案例辅助标签_周线五态": {"多头趋势": "C6EFCE", "突破启动": "F8CBAD",
                              "回踩修复": "FFF2CC", "均线蓄势": "DDEBF7",
                              "弱势结构": "FFC7CE", "未知": "E7E6E6"},
    "触发状态": {"新触发": "C6EFCE", "持续": "DDEBF7"},
}
PCT = ("距一年低点涨幅", "近120日收益", "MA20持续度", "距一年高点价格差",
       "平台深度")
NUM3 = ("平台缩量比", "平台收敛比", "R09核心质量分")
THIN = Border(*[Side("thin", color="D9D9D9")] * 4)
TOPB = Border(left=Side("thin", color="D9D9D9"), right=Side("thin", color="D9D9D9"),
              bottom=Side("thin", color="D9D9D9"), top=Side("medium", color="C00000"))


def write(ws, df, freeze="A2"):
    for c, name in enumerate(df.columns, 1):
        x = ws.cell(1, c, name)
        x.fill, x.font = NAVY, HF
        x.alignment = Alignment("center", "center", wrap_text=True)
    for i, (_, r) in enumerate(df.iterrows(), 2):
        for c, name in enumerate(df.columns, 1):
            v = r[name]
            cell = ws.cell(i, c, None if pd.isna(v) else v)
            cell.border = THIN
            cell.alignment = Alignment("center", "center")
            if name in PCT:
                cell.number_format = "0.0%"
            elif name in NUM3:
                cell.number_format = "0.000"
            elif name == "收盘价":
                cell.number_format = "0.00"
            elif name in ("RPS50", "RPS60", "RPS250"):
                cell.number_format = "0.0"
    for c, name in enumerate(df.columns, 1):
        w = max(len(str(name)) * 2.1,
                *(len(str(x)) * 1.15 for x in df[name].head(400).fillna("")))
        ws.column_dimensions[get_column_letter(c)].width = min(max(w, 8), 26)
    ws.freeze_panes = freeze


def main():
    src = f"{OUT}/stock_history_{CODE}.csv"
    d = pd.read_csv(src, encoding="utf-8-sig")
    d["观察日期"] = pd.to_datetime(d["观察日期"]).dt.date
    name = str(d["股票名称"].iloc[0])
    wb = Workbook()
    wb.remove(wb.active)

    # ---- 逐日 ----
    ws = wb.create_sheet("逐日观察")
    write(ws, d)
    u = d["统一信号"].fillna(0).astype(int).to_numpy()
    cols = list(d.columns)
    for i in range(len(d)):
        rr = i + 2
        if u[i] == 1:
            for c in range(1, len(cols) + 1):
                if not ws.cell(rr, c).fill.fgColor.rgb or \
                        ws.cell(rr, c).fill.fgColor.rgb == "00000000":
                    ws.cell(rr, c).fill = ROWSIG
            if i == 0 or u[i - 1] == 0:          # 段首:加粗红上边框
                for c in range(1, len(cols) + 1):
                    ws.cell(rr, c).border = TOPB
        for cn, mp in FILL.items():
            if cn not in cols:
                continue
            k = cols.index(cn) + 1
            f = mp.get(str(d.iloc[i][cn]))
            if f:
                ws.cell(rr, k).fill = PatternFill("solid", fgColor=f)
                if cn == "信号类型" and str(d.iloc[i][cn]) == "强确认":
                    ws.cell(rr, k).font = Font(bold=True, color="9C0006")

    # ---- 信号段 ----
    m = d["信号类型"].isin(["强确认", "标准确认"])
    seg, grp = [], (m != m.shift()).cumsum()
    for _, g in d.groupby(grp):
        if not bool(g["信号类型"].isin(["强确认", "标准确认"]).iloc[0]):
            continue
        p0, p1 = float(g["收盘价"].iloc[0]), float(g["收盘价"].iloc[-1])
        seg.append({"起": g["观察日期"].iloc[0], "止": g["观察日期"].iloc[-1],
                    "天数": len(g), "起收盘": p0, "止收盘": p1,
                    "段内涨跌": p1 / p0 - 1,
                    "段内最高": float(g["收盘价"].max()),
                    "段内最低": float(g["收盘价"].min()),
                    "强确认天数": int((g["信号类型"] == "强确认").sum()),
                    "标准确认天数": int((g["信号类型"] == "标准确认").sum()),
                    "段内出现平台信号":
                        "、".join(sorted(set(g.loc[g["平台信号"] != "无平台信号",
                                                   "平台信号"].astype(str)))) or "无"})
    sd = pd.DataFrame(seg)
    ws2 = wb.create_sheet("信号段")
    if len(sd):
        write(ws2, sd)
        for i in range(len(sd)):
            c = ws2.cell(i + 2, list(sd.columns).index("段内涨跌") + 1)
            c.number_format = "0.0%"
            c.fill = PatternFill("solid",
                                 fgColor="C6EFCE" if sd["段内涨跌"].iloc[i] > 0
                                 else "FFC7CE")
            c.font = Font(bold=True)
        for cn in ("起收盘", "止收盘", "段内最高", "段内最低"):
            k = list(sd.columns).index(cn) + 1
            for i in range(len(sd)):
                ws2.cell(i + 2, k).number_format = "0.00"

    # ---- 说明 ----
    ws3 = wb.create_sheet("说明", 0)
    lines = [
        f"{CODE} {name} 逐日观察数据 —— {d['观察日期'].iloc[0]} → {d['观察日期'].iloc[-1]}",
        f"共 {len(d)} 行(该股停牌/无行情的日子按最后有效价 ffill 参与,绝不剔除)",
        "",
        "■ 配色",
        "  整行浅红 = 统一信号=1(强确认或标准确认);段首有加粗红色上边框",
        "  信号类型:强确认 红(加粗)/ 标准确认 黄 / 观察级 蓝",
        "  平台信号:平台突破(研究)橙 / 平台观察 浅黄",
        "  触发状态:新触发 绿 / 持续 蓝",
        "  质量档:高 绿 / 低 红 / 缺失 灰   周线五态:多头趋势 绿 / 弱势结构 红",
        "",
        "■ 先看「信号段」页 —— 把统一信号=1 的连续段汇总成一行一段",
        "",
        "■ 口径",
        "  X01 分档用 RPS50(≥80 标准确认、≥90 强确认;三条件全中但 <80 为观察级)",
        "  平台强势日用 RPS50 ≥ 90(第一六八节统一)",
        "  分数列 R09/RPS 均为**全市场**横截面百分位,不是池内排名",
        "",
        "■ 必读的边界",
        "  平台筛选器三个部件单独检验一个都没过(选股 −5.27pp / p 0.810);",
        "  R08 是否决器不是选择器,R09 留出段十分位倒挂,净利增速与牛股概率呈 U 型。",
        "  本表是状态记录,不是买点,不构成任何投资建议。",
    ]
    for i, t in enumerate(lines, 1):
        c = ws3.cell(i, 1, t)
        if i == 1:
            c.fill, c.font = NAVY, Font(color="FFFFFF", bold=True, size=13)
        elif t.startswith("■"):
            c.font = Font(bold=True, color="17365D")
    ws3.cell(len(lines), 1).fill = PatternFill("solid", fgColor="FFC7CE")
    ws3.column_dimensions["A"].width = 96

    p = f"{OUT}/stock_history_{CODE}.xlsx"
    wb.save(p)
    print(f"已生成 {p}")
    print(f"  逐日 {len(d)} 行;信号段 {len(sd)} 段;"
          f"统一信号=1 共 {int(u.sum())} 天")


if __name__ == "__main__":
    main()
