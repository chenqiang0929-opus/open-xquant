"""§149 附:把每月 100 只清单导成带颜色分组的 Excel。

应用户要求
----------
「相对MA250 高30%、RPS60 高30%,帮我添加并且标注下颜色,
你觉得用处不大的字段放在后面,并且也标注下颜色」

列的分组与配色(依据是本节测出的**在选中样本内的剩余增量**)
------------------------------------------------------------
**绿色 · 选股条件**(规则本身用的两个,不可去掉)
    距一年低点涨幅、换手加速
**蓝色 · 有增量且训练/留出两段都稳**
    相对MA250(整体 lift 1.25;训练 1.17 / 留出 1.30)
    RPS60   (整体 1.15;训练 1.18 / 留出 1.14)
    并新增三个标记列:`相对MA250_高30%`、`RPS60_高30%`、`双高`
    (分位在**当月被选中的这批股票内部**计算,不是全市场分位)
**黄色 · 有信号但两段不稳,仅供参考**
    周线排列、周线已持续周、RPS120、120日收益率
**灰色 · 在已筛样本内无增量,放最后**
    RPS250(1.03)、MA20持续度120日(1.03)、20日波动率(0.97)、
    流通市值亿(1.03)、换手加速的剩余变异(0.94,已作门槛故无信息)

单元格高亮:`相对MA250` 与 `RPS60` 落在当月前 30% 的格子标绿;
双高的行在 `双高` 列标深绿。

**选股逻辑没有任何改动** —— 这些只是展示与标记。
**不构成任何买入建议。**
"""

from __future__ import annotations

import os

import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = os.environ.get("OXQ_OUT_DIR", "/home/user/oxq-panel")
SRC = f"{OUT}/startup_rule_top100_full_2019_2026.csv"
DST = f"{OUT}/startup_rule_top100.xlsx"

GRP = {
    "基础": (["观察日", "排名", "代码", "名称", "申万一级"], "FFFFFFFF", "FF000000"),
    "选股条件": (["距一年低点涨幅", "换手加速"], "FFC6EFCE", "FF006100"),
    "有增量·两段都稳": (["相对MA250", "相对MA250_高30%", "RPS60", "RPS60_高30%",
                    "双高"], "FFBDD7EE", "FF1F4E79"),
    "有信号·两段不稳": (["周线排列", "周线已持续周", "RPS120", "120日收益率"],
                   "FFFFE699", "FF7F6000"),
    "无增量·仅备查": (["RPS250", "MA20持续度120日", "20日波动率", "流通市值亿"],
                  "FFD9D9D9", "FF595959"),
    "结果": (["未来60日最大涨幅", "启动(>=50%)"], "FFF8CBAD", "FF833C00"),
}


def main():
    df = pd.read_csv(SRC)
    # 月内分位:在当月被选中的这批股票内部
    for col, flag in (("相对MA250", "相对MA250_高30%"), ("RPS60", "RPS60_高30%")):
        r = df.groupby("观察日")[col].rank(pct=True, ascending=True)
        df[flag] = (r >= 0.70).where(df[col].notna())
    df["双高"] = df["相对MA250_高30%"].fillna(False) & df["RPS60_高30%"].fillna(False)

    order = [c for _, (cs, _, _) in GRP.items() for c in cs]
    miss = [c for c in order if c not in df.columns]
    assert not miss, f"缺列 {miss}"
    df = df[order]

    # 历史效力:双高 vs 其余(只用已有结果的行)
    dd = df[df["启动(>=50%)"].notna()].copy()
    dd["y"] = dd["启动(>=50%)"].astype(bool)
    b = dd.y.mean()
    hh = dd[dd["双高"]]
    print(f"已有结果 {len(dd):,} 行,基准启动率 {b:.2%}")
    print(f"**双高** {len(hh):,} 行({len(hh)/len(dd):.1%}),"
          f"启动率 **{hh.y.mean():.2%}**,lift **{hh.y.mean()/b:.2f}**")
    dd["年"] = pd.to_datetime(dd.观察日).dt.year
    print("\n双高逐年:")
    for yy, g in dd.groupby("年"):
        h = g[g["双高"]]
        if len(h) < 20:
            continue
        print(f"  {yy}  双高 {len(h):>3d} 只  启动率 {h.y.mean():>6.2%}  "
              f"当年基准 {g.y.mean():>6.2%}  lift {h.y.mean()/g.y.mean():.2f}")

    with pd.ExcelWriter(DST, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="每月100只", freeze_panes=(2, 5))
        ws = w.sheets["每月100只"]
        ws.insert_rows(1)
        ci = 1
        for gname, (csx, fill, font) in GRP.items():
            a, bcol = get_column_letter(ci), get_column_letter(ci + len(csx) - 1)
            if len(csx) > 1:
                ws.merge_cells(f"{a}1:{bcol}1")
            c = ws[f"{a}1"]
            c.value = gname
            c.fill = PatternFill("solid", fgColor=fill)
            c.font = Font(bold=True, color=font)
            c.alignment = Alignment(horizontal="center")
            for k in range(len(csx)):
                h = ws.cell(row=2, column=ci + k)
                h.fill = PatternFill("solid", fgColor=fill)
                h.font = Font(bold=True, color=font)
            ci += len(csx)
        n = len(df) + 2
        cols = {c: i + 1 for i, c in enumerate(df.columns)}
        grn = PatternFill("solid", fgColor="FFC6EFCE")
        dgrn = PatternFill("solid", fgColor="FF70AD47")
        for c in ("相对MA250_高30%", "RPS60_高30%"):
            lt = get_column_letter(cols[c])
            ws.conditional_formatting.add(
                f"{lt}3:{lt}{n}",
                CellIsRule(operator="equal", formula=['TRUE'], fill=grn))
        lt = get_column_letter(cols["双高"])
        ws.conditional_formatting.add(
            f"{lt}3:{lt}{n}",
            CellIsRule(operator="equal", formula=['TRUE'], fill=dgrn,
                       font=Font(bold=True, color="FFFFFFFF")))
        lt = get_column_letter(cols["启动(>=50%)"])
        ws.conditional_formatting.add(
            f"{lt}3:{lt}{n}",
            CellIsRule(operator="equal", formula=['TRUE'],
                       fill=PatternFill("solid", fgColor="FFFFC000")))
        for c, i in cols.items():
            ws.column_dimensions[get_column_letter(i)].width = \
                max(9, min(16, len(str(c)) * 1.6 + 3))
        ws.auto_filter.ref = f"A2:{get_column_letter(len(df.columns))}{n}"
    print(f"\n落库 {DST}({len(df):,} 行 × {len(df.columns)} 列)")


if __name__ == "__main__":
    main()
