"""§156 交付:把逐年买点清单与实际成交流水打包成 Excel(不做任何计算,只排版)。"""

from __future__ import annotations

import os

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = os.environ.get("OXQ_OUT_DIR", "/home/user/oxq-panel")
HDR = PatternFill("solid", fgColor="17365D")
HF = Font(color="FFFFFF", bold=True, size=10)
GOOD = PatternFill("solid", fgColor="C6EFCE")
WARN = PatternFill("solid", fgColor="FFF2CC")
BAD = PatternFill("solid", fgColor="FFC7CE")


def put(ws, df, pcts=(), color=None):
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        ws.cell(1, c).fill = HDR
        ws.cell(1, c).font = HF
        ws.cell(1, c).alignment = Alignment("center", "center", wrap_text=True)
    for _, r in df.iterrows():
        ws.append([None if (isinstance(v, float) and not np.isfinite(v)) else v
                   for v in r.tolist()])
    for i, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(
            9, min(15, len(str(col)) + 4))
        if col in pcts:
            for row in range(2, len(df) + 2):
                ws.cell(row, i).number_format = "0.0%"
    if color and color in df.columns:
        k = list(df.columns).index(color) + 1
        for row in range(2, len(df) + 2):
            v = ws.cell(row, k).value
            if isinstance(v, (int, float)) and np.isfinite(v):
                ws.cell(row, k).fill = GOOD if v > 0.10 else (
                    BAD if v < -0.05 else WARN)
    ws.freeze_panes = "A2"


def main():
    a = pd.read_csv(f"{OUT}/platform_buypoints.csv", dtype={"代码": str})
    b = pd.read_csv(f"{OUT}/platform_trades.csv", dtype={"代码": str})
    s = pd.read_csv(f"{OUT}/platform_yearly_summary.csv")
    a["代码"] = a["代码"].str.zfill(6)
    b["代码"] = b["代码"].str.zfill(6)
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    lines = [
        ("平台筛选器 2016–2026 逐年买点清单与实际成交流水", 0),
        (f"买点 {len(a):,} 个,涉及 {a['代码'].nunique():,} 只;"
         f"框架实际成交 {len(b):,} 笔", 1),
        ("", 0),
        ("■ 先读这一条:这不是选股模型", 0),
        ("第六十一节:三条全中组合年化 +10.37%,但 300 次随机对照 p = 0.16,"
         "与随机无法区分。", 1),
        ("第一五五节:接上买点+止损+择时后年化 +15.83%、回撤 16.7%,但同市值同行业"
         "随机对照拿到 +18.29%,超额 −2.46pp、p = 0.656 —— 不通过。", 1),
        ("→ 这套东西的价值在框架(在波动收敛处买、止损放结构位、大盘不好空仓),"
         "不在「买哪一只」。本表是观察名单,不是买入指令。", 1),
        ("", 0),
        ("■ 两张表的关系:先看「实际成交」,再看逐年买点", 0),
        (f"逐年买点表一共 {len(a):,} 行,是「所有亮灯并突破的时刻」,看不过来;", 1),
        (f"「实际成交」{len(b):,} 笔才是 10 槽位框架真正会买的东西。", 1),
        ("", 0),
        ("■ 三个必须知道的事实", 0),
        (f"1) 卖出原因里 {int((b['卖出原因']=='大盘过滤清仓').sum())} 笔是「大盘过滤清仓」"
         f"(占 {(b['卖出原因']=='大盘过滤清仓').mean():.0%}),止损只有 "
         f"{int((b['卖出原因']=='止损').sum())} 笔 —— 真正在管风险的是大盘开关,不是止损。", 1),
        (f"2) 全部买点的「后 60 日收益中位数」逐年多为负;整体胜率只有 "
         f"{(b['收益']>0).mean():.1%},单笔中位 {b['收益'].median():+.2%},"
         f"但单笔均值 {b['收益'].mean():+.2%} —— 钱全来自右尾少数几笔。", 1),
        ("3) 2018 年成交 0 笔,因为大盘过滤全年关闭。这是过滤器按设计工作,不是数据缺失。", 1),
        ("", 0),
        ("■ 口径", 0),
        ("绝对阈值(legacy):缩量比<0.80、收敛比<0.80、深度≤0.352,调整天数≥15,"
         "要求 20 周线向上", 1),
        ("买点:三条全中 且 收盘 > 平台内(强势日→前一日)最高收盘", 1),
        ("止损:平台下沿(平台内→前一日最低收盘);距买入价超 15% 则上移到 −15%", 1),
        ("大盘过滤:全市场等权净值 < 自身 MA200 → 清仓且不新开仓", 1),
        ("仓位:10 等权槽位,突破日先到先得,同日按收敛比升序;持有上限 120 交易日", 1),
        ("价格 ffill 参与,退市股绝不剔除", 1),
        ("", 0),
        ("注意:候选筛选与后验描述,不构成投资建议。仓位、滑点、涨跌停、流动性均未纳入。", 0)]
    for txt, i in lines:
        ws.append([("    " * i) + txt])
    ws["A1"].fill = HDR
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    for r in (4, 9, 13, 19):
        ws.cell(r, 1).font = Font(bold=True, color="17365D")
    ws.cell(len(lines), 1).fill = BAD
    ws.column_dimensions["A"].width = 105

    put(wb.create_sheet("逐年汇总"), s,
        pcts=("大盘开占比", "后60日中位", "60日峰值ge50", "触止损占比", "胜率",
              "单笔中位", "单笔均值"), color="单笔均值")
    bt = b[["买入日", "卖出日", "持有交易日", "代码", "名称", "申万一级", "买入价",
            "卖出价", "止损价", "收益", "卖出原因", "年"]].sort_values(
        "收益", ascending=False)
    put(wb.create_sheet(f"实际成交{len(b)}笔"), bt, pcts=("收益",), color="收益")
    cols = ["买点日", "代码", "名称", "申万一级", "调整天数", "深度", "缩量比",
            "收敛比", "买入价", "止损价", "止损距离", "止损σ倍数", "大盘过滤开",
            "流通市值亿", "后20日", "后60日", "后120日", "后250日",
            "60日内峰值涨幅", "60日内触及止损"]
    pcts = ("深度", "止损距离", "后20日", "后60日", "后120日", "后250日",
            "60日内峰值涨幅")
    for y in sorted(a["年"].unique()):
        d = a[a["年"] == y][cols].sort_values("收敛比")
        put(wb.create_sheet(f"{y}买点{len(d)}"), d, pcts=pcts, color="后60日")
    f = f"{OUT}/platform_yearly_lists.xlsx"
    wb.save(f)
    print(f"已生成 {f}(工作表 {len(wb.sheetnames)} 张,买点 {len(a):,} 行,"
          f"成交 {len(b):,} 行)")


if __name__ == "__main__":
    main()
