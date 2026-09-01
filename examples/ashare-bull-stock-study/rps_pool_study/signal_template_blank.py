"""生成「价格信号 + 平台信号 + 质量因子」空白监控模板(供 Codex 与本项目共用)。

**本脚本不做任何计算,只生成一个空模板** —— 表头、分组色带、字段字典、
配色图例、以及**会自动勾稽的检查页与恒等式自检页**。

相对 Codex v0.4 工作簿补的两处(他那版的检查页只数字段个数,不查值):
1. **检查页改成真公式**:粘入数据后自动统计并与勾稽项比对,差异非 0 会显示「✗」;
2. **新增恒等式自检页**:逐条检查「统一信号 ⟺ 强确认或标准确认」
   「触发状态 ⟺ 连续确认天数」「周线多头排列 ⟺ 五态=多头趋势」
   「强确认 ⟺ RPS60≥90 且三条件」等值域恒等式,违例数必须为 0;
3. **新增数据锚点页**:面板行列数、日期范围、规则哈希、数据快照哈希 ——
   规则可追溯之外,数据也要可追溯。
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.environ.get("OXQ_OUT_DIR", "/home/user/oxq-panel")
NAVY = PatternFill("solid", fgColor="17365D")
HF = Font(color="FFFFFF", bold=True, size=10)
BAND = {"基础与样本": "595959", "信号输出": "1F6FC5", "平台与质量": "1F7A5C",
        "价格与趋势": "8A6D1F"}
GRP = [("基础与样本", ["样本类型", "观察日期", "股票代码", "股票名称", "收盘价"]),
       ("信号输出", ["统一信号", "信号类型", "信号理由", "首次触发日期",
                  "连续确认天数", "触发状态"]),
       ("平台与质量", ["平台信号", "周线多头排列", "案例展示分层_质量",
                   "案例辅助标签_周线五态", "平台深度", "平台缩量比", "平台收敛比",
                   "R09核心质量分"]),
       ("价格与趋势", ["距一年低点涨幅", "近120日收益", "MA20持续度", "RPS60",
                   "RPS250", "距一年高点价格差"])]
COLS = [c for _, cs in GRP for c in cs]
NC = len(COLS)
CI = {c: i for i, c in enumerate(COLS, 1)}          # 列号
CL_ = {c: get_column_letter(i) for c, i in CI.items()}
PCT = ("距一年低点涨幅", "近120日收益", "MA20持续度", "距一年高点价格差", "平台深度")
NUM2 = ("RPS60", "RPS250", "R09核心质量分", "平台缩量比", "平台收敛比", "收盘价")
LEGEND = [
    ("信号类型", [("强确认", "FFC7CE"), ("标准确认", "FFEB9C"), ("观察级", "DDEBF7"),
              ("无信号", "FFFFFF")]),
    ("平台信号", [("平台突破（研究）", "F8CBAD"), ("平台观察", "FFF2CC"),
              ("无平台信号", "FFFFFF")]),
    ("案例展示分层_质量", [("高", "C6EFCE"), ("中", "FFFFFF"), ("低", "FFC7CE"),
                   ("缺失", "E7E6E6")]),
    ("案例辅助标签_周线五态", [("多头趋势", "C6EFCE"), ("突破启动", "F8CBAD"),
                      ("回踩修复", "FFF2CC"), ("均线蓄势", "DDEBF7"),
                      ("弱势结构", "FFC7CE"), ("未知", "E7E6E6")]),
    ("触发状态", [("新触发", "C6EFCE"), ("持续", "DDEBF7"), ("未触发", "FFFFFF")]),
]
THIN = Border(*[Side("thin", color="D9D9D9")] * 4)
DATA = "全部清单"
R0 = 3          # 数据起始行(1=色带 2=列名 3=第一行数据)


def head(ws, note=None):
    """两层表头:分组色带 + 列名;返回数据起始行。"""
    i = 1
    for g, cs in GRP:
        ws.merge_cells(start_row=1, start_column=i, end_row=1,
                       end_column=i + len(cs) - 1)
        c = ws.cell(1, i, g)
        c.fill = PatternFill("solid", fgColor=BAND[g])
        c.font = HF
        c.alignment = Alignment("center", "center")
        i += len(cs)
    for k, col in enumerate(COLS, 1):
        c = ws.cell(2, k, col)
        c.fill = NAVY
        c.font = HF
        c.alignment = Alignment("center", "center", wrap_text=True)
        ws.column_dimensions[get_column_letter(k)].width = max(10, min(18,
                                                                      len(col) + 4))
    for k, col in enumerate(COLS, 1):
        f = "0.0%" if col in PCT else ("0.00" if col in NUM2 else None)
        if f:
            for r in range(R0, R0 + 2000):
                ws.cell(r, k).number_format = f
        for r in range(R0, R0 + 2000):
            ws.cell(r, k).border = THIN
    ws.freeze_panes = ws.cell(R0, 5)
    ws.auto_filter.ref = f"A2:{get_column_letter(NC)}2"
    if note:
        ws.cell(1, NC + 2, note).font = Font(bold=True, color="C00000")
    return R0


DICT = [
    ("基础与样本", "样本类型", "sample_type", "月末观察 / 平台突破日 / 其他自定义",
     "生成器", "正式", "不得为空"),
    ("基础与样本", "观察日期", "obs_date", "交易日;所有字段一律不得晚于该日",
     "行情面板", "正式", "不得为空"),
    ("基础与样本", "股票代码", "code", "六位,不含交易所前缀,左补零",
     "行情面板", "正式", "不得为空"),
    ("基础与样本", "股票名称", "name", "观察日当时名称;缺失留空不影响信号",
     "名称表", "展示", "留空"),
    ("基础与样本", "收盘价", "close", "前复权收盘;前复权锚点不同只影响绝对值,不影响比值",
     "行情面板", "正式", "缺失则整行不具备资格"),
    ("信号输出", "统一信号", "unified_signal",
     "标准确认或强确认记 1,观察级与无信号记 0,行情缺失记 NA", "X01 v1.0", "正式", "NA"),
    ("信号输出", "信号类型", "signal_type",
     "无信号 / 观察级 / 标准确认 / 强确认;三条件全中且 RPS60<80 = 观察级,"
     "≥80 = 标准确认,≥90 = 强确认", "X01 v1.0", "正式", "无信号"),
    ("信号输出", "信号理由", "signal_reason", "列出实际触发的条件",
     "X01 v1.0", "正式", "留空"),
    ("信号输出", "首次触发日期", "first_trigger_date",
     "当前这段连续「统一信号=1」的起始交易日", "生成器", "正式", "未触发则空"),
    ("信号输出", "连续确认天数", "consecutive_days",
     "截至观察日,统一信号连续为 1 的交易日数;未触发为 0", "生成器", "正式", "0"),
    ("信号输出", "触发状态", "trigger_state",
     "新触发(连续=1) / 持续(连续>1) / 未触发(连续=0)", "生成器", "正式", "未触发"),
    ("平台与质量", "平台信号", "platform_signal",
     "平台突破（研究）= 三条全中且收盘严格突破平台上沿;平台观察 = 三条全中;其余无平台信号",
     "平台规则", "研究(WATCHLIST)", "无平台信号"),
    ("平台与质量", "周线多头排列", "weekly_bull_alignment_flag",
     "周收盘 > MA20周 > MA60周;二元,不改写正式价格信号", "X01 v1.0 辅助", "正式", "FALSE"),
    ("平台与质量", "案例展示分层_质量", "quality_tier_display",
     "核心质量分 ≥0.70 高 / 0.30~0.70 中 / <0.30 低 / 缺失;"
     "**固定数值切点,仅展示,不是正式 R09 门槛**", "展示分层", "展示", "缺失"),
    ("平台与质量", "案例辅助标签_周线五态", "weekly_state_label",
     "有序判别,先命中优先:未知 → 多头趋势(周收>MA20周>MA60周) → "
     "突破启动(周收≥MA20周 且 日线ret_5>5%) → 回踩修复(周收≥0.95×MA60周 且 周收<MA20周) "
     "→ 均线蓄势(|周收/MA20周−1|≤5%) → 弱势结构。**辅助标签,非正式规则**",
     "案例生成器", "展示", "未知"),
    ("平台与质量", "平台深度", "platform_depth",
     "1 − min(low[ts:t]) / max(high[ts:t]);ts = 最近强势日,窗口含 ts 与观察日",
     "平台规则", "研究", "空"),
    ("平台与质量", "平台缩量比", "platform_volume_ratio",
     "mean(volume[ts:t]) / mean(volume[ts−60:ts−1]);**成交量用原始值,不做复权反调整**",
     "平台规则", "研究", "空"),
    ("平台与质量", "平台收敛比", "platform_atr_ratio",
     "mean(TR[td:t]) / mean(TR[ts−60:ts−1]);TR 为真实波幅(含跳空);"
     "td = ts 之后首个 low≤MA100×1.03 **且 MA100 向上**(MA100[k]>MA100[k−20]) 的日子",
     "平台规则", "研究", "空"),
    ("平台与质量", "R09核心质量分", "r09_core_quality_score",
     "四因子(TTM净利率>0 / ROE>0 / ep_ttm>0 / ep&cfp>0且转换率>0)各自 eligibility → "
     "有效横截面 1%/99% 缩尾 → 升序百分位 → 等权平均;**任一缺失则整体缺失**",
     "R09 正式", "正式", "缺失"),
    ("价格与趋势", "距一年低点涨幅", "recovery_from_low_250",
     "收盘 / 过去250个交易日最低收盘 − 1(含当日)", "X01 v1.0", "正式", "不具备资格"),
    ("价格与趋势", "近120日收益", "ret_120", "收盘 / 120交易日前收盘 − 1",
     "X01 v1.0", "正式", "不具备资格"),
    ("价格与趋势", "MA20持续度", "above_ma20_share_120",
     "过去120个交易日中收盘 > 当日MA20 的比例", "X01 v1.0", "正式", "不具备资格"),
    ("价格与趋势", "RPS60", "rps_60",
     "60日收益在**策略池当日收益可算者**中的百分位×100;排名步骤不逐日剔除"
     "ST/停牌/零成交;最低横截面样本 100", "RPS 缓存", "正式", "不具备资格"),
    ("价格与趋势", "RPS250", "rps_250", "同上,250日窗口", "RPS 缓存", "正式", "空"),
    ("价格与趋势", "距一年高点价格差", "gap_to_high_250",
     "收盘 / 过去250个交易日最高收盘 − 1(≤0)", "X01 v1.0 辅助", "正式", "空"),
]


def main():  # noqa: PLR0915
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    lines = [
        ("价格信号 + 平台信号 + 质量因子 —— 空白监控模板 v1.0", 0),
        ("把数据粘进「全部清单」页,其余各页(检查/恒等式自检/信号交叉表/逐月汇总)"
         "会自动算出来。", 1),
        ("", 0),
        ("■ 填表顺序", 0),
        ("1) 先填「数据锚点」页 —— 面板行列数、日期范围、规则哈希、数据快照哈希。"
         "规则可追溯之外,数据也必须可追溯:同一规则在不同数据快照上会出不同结果。", 1),
        ("2) 把逐行数据粘进「全部清单」页第 3 行起,列顺序不要动。", 1),
        ("3) 看「检查」页:每一行的「差异」必须为 0,状态列显示 ✓。", 1),
        ("4) 看「恒等式自检」页:每条违例数必须为 0。**这一页才是真正能抓错的地方。**", 1),
        ("5) 子集页(重点·双信号 / 强确认 / 平台突破日)按需筛选后另存,或直接用自动筛选。", 1),
        ("", 0),
        ("■ 与常见做法的两处差别(建议保留)", 0),
        ("A. 检查页用**真公式**自动勾稽,不是手填的字段计数。"
         "只数「字段有没有」查不出任何值错误。", 1),
        ("B. 新增**恒等式自检**页。逐条检查值域恒等式,例如"
         "「统一信号=1 ⟺ 信号类型∈{强确认,标准确认}」。这类检查一行公式就能写,而且真会抓到错。", 1),
        ("", 0),
        ("■ 配色规则(静态填充,不用条件格式)", 0),
        ("条件格式在 WPS 里会出现白字白底看不见的情况,所以本模板一律用静态填充。", 1),
        ("", 0),
        ("■ 四级验证状态(建议取代散乱标签)", 0),
        ("正式    —— 已通过样本外 + 随机对照,可参与信号", 1),
        ("已实现  —— 口径已冻结、实现完成,但未做样本外或未做对照", 1),
        ("研究    —— WATCHLIST,只作事件池与状态标记,不得当买点", 1),
        ("展示    —— 仅为读表方便而加的分层/标签,不是任何规则的一部分", 1),
        ("", 0),
        ("■ 边界(必须保留在交付物里)", 0),
        ("平台信号为研究状态,不是买点:已知存在假突破(突破日 RPS 已回落、"
         "收盘位于当日振幅偏低、长上影、次日跌回上沿),尚缺当日RPS/K线质量/次日确认等过滤。", 1),
        ("「当日收盘确认 + 当日收盘成交」在实盘不可执行,回测须改为次日开盘或次日收盘。", 1),
        ("这是候选筛选与状态标记,不是买入指令,不构成投资建议。", 0)]
    for t, i in lines:
        ws.append([("    " * i) + t])
    ws["A1"].fill = NAVY
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    for r in (4, 11, 15, 18, 24):
        ws.cell(r, 1).font = Font(bold=True, color="17365D")
    ws.cell(len(lines), 1).fill = PatternFill("solid", fgColor="FFC7CE")
    ws.column_dimensions["A"].width = 110
    r = len(lines) + 2
    ws.cell(r, 1, "配色图例").font = Font(bold=True, color="17365D")
    r += 1
    for col, items in LEGEND:
        ws.cell(r, 1, col).font = Font(bold=True)
        for k, (v, f) in enumerate(items, 2):
            c = ws.cell(r, k, v)
            c.fill = PatternFill("solid", fgColor=f)
            c.border = THIN
            ws.column_dimensions[get_column_letter(k)].width = 16
        r += 1

    # ---- 字段字典 ----
    wd = wb.create_sheet("字段字典")
    hdr = ["模块", "中文标题", "字段名", "定义/计算口径", "来源", "验证状态", "缺失处理"]
    wd.append(hdr)
    for k in range(1, len(hdr) + 1):
        wd.cell(1, k).fill = NAVY
        wd.cell(1, k).font = HF
        wd.cell(1, k).alignment = Alignment("center", "center")
    for row in DICT:
        wd.append(list(row))
    for k, wdt in enumerate((12, 20, 26, 62, 14, 16, 20), 1):
        wd.column_dimensions[get_column_letter(k)].width = wdt
    for r in range(2, len(DICT) + 2):
        wd.cell(r, 4).alignment = Alignment(wrap_text=True, vertical="top")
        st = wd.cell(r, 6).value
        wd.cell(r, 6).fill = PatternFill("solid", fgColor={
            "正式": "C6EFCE", "已实现": "DDEBF7",
            "研究(WATCHLIST)": "FFF2CC", "研究": "FFF2CC",
            "展示": "E7E6E6"}.get(st, "FFFFFF"))
        for k in range(1, len(hdr) + 1):
            wd.cell(r, k).border = THIN
    wd.freeze_panes = "A2"

    # ---- 数据锚点(填写页)----
    wa = wb.create_sheet("数据锚点")
    wa.append(["锚点项", "填写值", "说明"])
    for k in range(1, 4):
        wa.cell(1, k).fill = NAVY
        wa.cell(1, k).font = HF
    for a, b in [("面板交易日数", "同一规则在不同数据快照上会出不同结果,必须记下来"),
                 ("面板股票只数", "含退市股;若剔除过退市股请在此注明"),
                 ("面板首日", ""), ("面板末日", "与他方核对时先看这一项"),
                 ("规则文件与版本", "如 x01_price_start_v1.0.0.json"),
                 ("规则哈希", ""), ("数据快照哈希", "行情/财务缓存的内容哈希"),
                 ("复权方式", "前复权/后复权;成交量是否随之调整(建议:不调整)"),
                 ("RPS 横截面池", "池内只数与是否逐日剔除 ST/停牌/零成交"),
                 ("财务生效规则", "如:公告日之后第一个交易日生效"),
                 ("生成时间", ""), ("生成脚本", "")]:
        wa.append([a, "", b])
    for k, wdt in enumerate((22, 34, 56), 1):
        wa.column_dimensions[get_column_letter(k)].width = wdt
    for r in range(2, 14):
        wa.cell(r, 2).fill = PatternFill("solid", fgColor="FFF2CC")
        for k in range(1, 4):
            wa.cell(r, k).border = THIN

    # ---- 数据页 ----
    head(wb.create_sheet(DATA), "← 数据从第 3 行起粘贴,列顺序不要改")
    for nm in ("重点·双信号", "强确认", "平台突破日", "锚点对照(可选)"):
        head(wb.create_sheet(nm), "← 从「全部清单」筛选后粘入")

    # ---- 检查页(真公式)----
    wc = wb.create_sheet("检查")
    wc.append(["检查项", "实际值", "勾稽值", "差异", "状态", "说明"])
    for k in range(1, 7):
        wc.cell(1, k).fill = NAVY
        wc.cell(1, k).font = HF
    rng = f"'{DATA}'!{{c}}3:{{c}}60000"
    g, f_, t_ = CL_["信号类型"], CL_["统一信号"], CL_["平台信号"]
    q, w5c, wb_ = CL_["案例展示分层_质量"], CL_["案例辅助标签_周线五态"], CL_["周线多头排列"]
    checks = [
        ("总行数", f"=COUNTA({rng.format(c=CL_['股票代码'])})", "", "手工填入预期行数"),
        ("涉及股票只数",
         f"=SUMPRODUCT(({rng.format(c=CL_['股票代码'])}<>\"\")/"
         f"COUNTIF({rng.format(c=CL_['股票代码'])},{rng.format(c=CL_['股票代码'])}&\"\"))",
         "", "去重计数"),
        ("强确认", f"=COUNTIF({rng.format(c=g)},\"强确认\")", "", ""),
        ("标准确认", f"=COUNTIF({rng.format(c=g)},\"标准确认\")", "", ""),
        ("观察级", f"=COUNTIF({rng.format(c=g)},\"观察级\")", "", ""),
        ("无信号", f"=COUNTIF({rng.format(c=g)},\"无信号\")", "", ""),
        ("统一信号=1", f"=SUMIF({rng.format(c=f_)},1)",
         "=B4+B5", "**必须等于 强确认+标准确认**"),
        ("平台突破（研究）", f"=COUNTIF({rng.format(c=t_)},\"平台突破（研究）\")", "", ""),
        ("平台观察", f"=COUNTIF({rng.format(c=t_)},\"平台观察\")", "", ""),
        ("无平台信号", f"=COUNTIF({rng.format(c=t_)},\"无平台信号\")", "", ""),
        ("平台信号三档合计", "=B9+B10+B11", "=B2", "必须等于总行数"),
        ("质量:高/中/低/缺失合计",
         f"=COUNTIF({rng.format(c=q)},\"高\")+COUNTIF({rng.format(c=q)},\"中\")"
         f"+COUNTIF({rng.format(c=q)},\"低\")+COUNTIF({rng.format(c=q)},\"缺失\")",
         "=B2", "必须等于总行数"),
        ("周线五态合计",
         "+".join(f"COUNTIF({rng.format(c=w5c)},\"{v}\")"
                  for v in ("多头趋势", "突破启动", "回踩修复", "均线蓄势",
                            "弱势结构", "未知")).join(("=", "")),
         "=B2", "必须等于总行数"),
        ("周线多头排列为真", f"=COUNTIF({rng.format(c=wb_)},TRUE)",
         f"=COUNTIF({rng.format(c=w5c)},\"多头趋势\")",
         "**必须等于 五态=多头趋势 的行数**"),
    ]
    for i, (a, b, cc, note) in enumerate(checks, 2):
        wc.cell(i, 1, a)
        wc.cell(i, 2, b)
        if cc:
            wc.cell(i, 3, cc)
            wc.cell(i, 4, f"=B{i}-C{i}")
            wc.cell(i, 5, f'=IF(D{i}=0,"✓","✗")')
        wc.cell(i, 6, note)
    for k, wdt in enumerate((26, 14, 14, 10, 8, 40), 1):
        wc.column_dimensions[get_column_letter(k)].width = wdt
    for r in range(2, len(checks) + 2):
        for k in range(1, 7):
            wc.cell(r, k).border = THIN
        if wc.cell(r, 3).value:
            wc.cell(r, 5).fill = PatternFill("solid", fgColor="FFF2CC")

    # ---- 恒等式自检(真正能抓错的一页)----
    wi = wb.create_sheet("恒等式自检")
    wi.append(["恒等式", "违例数", "状态", "为什么要查"])
    for k in range(1, 5):
        wi.cell(1, k).fill = NAVY
        wi.cell(1, k).font = HF
    rec, r120, mf = CL_["距一年低点涨幅"], CL_["近120日收益"], CL_["MA20持续度"]
    rp, run_, st = CL_["RPS60"], CL_["连续确认天数"], CL_["触发状态"]
    ft, od = CL_["首次触发日期"], CL_["观察日期"]

    def rr(c):
        return f"'{DATA}'!{c}3:{c}60000"
    ids = [
        ("统一信号=1 ⟺ 信号类型∈{强确认,标准确认}",
         f'=SUMPRODUCT(({rr(f_)}=1)*({rr(g)}<>"强确认")*({rr(g)}<>"标准确认"))'
         f'+SUMPRODUCT(({rr(f_)}=0)*(({rr(g)}="强确认")+({rr(g)}="标准确认")))',
         "两列各自生成时最容易脱钩"),
        ("强确认 ⟹ RPS60≥90",
         f'=SUMPRODUCT(({rr(g)}="强确认")*({rr(rp)}<90))',
         "分档门槛写错会在这里暴露"),
        ("标准确认 ⟹ 80≤RPS60<90",
         f'=SUMPRODUCT(({rr(g)}="标准确认")*(({rr(rp)}<80)+({rr(rp)}>=90)))', ""),
        ("观察级 ⟹ RPS60<80",
         f'=SUMPRODUCT(({rr(g)}="观察级")*({rr(rp)}>=80))', ""),
        ("信号类型≠无信号 ⟹ 三条件全中",
         f'=SUMPRODUCT(({rr(g)}<>"无信号")*({rr(g)}<>"")*'
         f'(({rr(rec)}<0.4)+({rr(r120)}<0.1)+({rr(mf)}<0.55)))',
         "三个观察条件与分档之间的一致性"),
        ("触发状态=新触发 ⟺ 连续确认天数=1",
         f'=SUMPRODUCT(({rr(st)}="新触发")*({rr(run_)}<>1))'
         f'+SUMPRODUCT(({rr(st)}<>"新触发")*({rr(run_)}=1))', ""),
        ("触发状态=未触发 ⟺ 连续确认天数=0",
         f'=SUMPRODUCT(({rr(st)}="未触发")*({rr(run_)}<>0))'
         f'+SUMPRODUCT(({rr(st)}<>"未触发")*({rr(run_)}=0))', ""),
        ("连续确认天数>0 ⟹ 首次触发日期非空 且 ≤观察日",
         f'=SUMPRODUCT(({rr(run_)}>0)*(({rr(ft)}="")+({rr(ft)}>{rr(od)})))',
         "**首次触发日期晚于观察日 = 前视**"),
        ("周线多头排列=TRUE ⟺ 五态=多头趋势",
         f'=SUMPRODUCT(({rr(wb_)}=TRUE)*({rr(w5c)}<>"多头趋势"))'
         f'+SUMPRODUCT(({rr(wb_)}<>TRUE)*({rr(w5c)}="多头趋势"))',
         "正式二元字段与展示标签必须自洽"),
        ("平台深度 ∈ [0,1)",
         f'=SUMPRODUCT(({rr(CL_["平台深度"])}<>"")*'
         f'(({rr(CL_["平台深度"])}<0)+({rr(CL_["平台深度"])}>=1)))', ""),
        ("平台信号≠无平台信号 ⟹ 平台三值齐全",
         f'=SUMPRODUCT(({rr(t_)}<>"无平台信号")*({rr(t_)}<>"")*'
         f'(({rr(CL_["平台深度"])}="")+({rr(CL_["平台缩量比"])}="")'
         f'+({rr(CL_["平台收敛比"])}="")))', "三条全中却算不出三值 = 生成器脱钩"),
        ("质量分∈[0,1] 且 与分层自洽",
         f'=SUMPRODUCT(({rr(CL_["R09核心质量分"])}<>"")*'
         f'(({rr(CL_["R09核心质量分"])}<0)+({rr(CL_["R09核心质量分"])}>1)))'
         f'+SUMPRODUCT(({rr(q)}="高")*({rr(CL_["R09核心质量分"])}<0.7))'
         f'+SUMPRODUCT(({rr(q)}="低")*({rr(CL_["R09核心质量分"])}>=0.3))', ""),
        ("距一年高点价格差 ≤ 0",
         f'=SUMPRODUCT(({rr(CL_["距一年高点价格差"])}<>"")*'
         f'({rr(CL_["距一年高点价格差"])}>0.000001))',
         "现价不可能高于含当日的 250 日最高"),
        ("MA20持续度 ∈ [0,1]",
         f'=SUMPRODUCT(({rr(mf)}<>"")*(({rr(mf)}<0)+({rr(mf)}>1)))', ""),
        ("RPS60 / RPS250 ∈ [0,100]",
         f'=SUMPRODUCT(({rr(rp)}<>"")*(({rr(rp)}<0)+({rr(rp)}>100)))'
         f'+SUMPRODUCT(({rr(CL_["RPS250"])}<>"")*'
         f'(({rr(CL_["RPS250"])}<0)+({rr(CL_["RPS250"])}>100)))', ""),
    ]
    for i, (a, b, note) in enumerate(ids, 2):
        wi.cell(i, 1, a)
        wi.cell(i, 2, b)
        wi.cell(i, 3, f'=IF(B{i}=0,"✓","✗ 有违例")')
        wi.cell(i, 4, note)
        wi.cell(i, 3).fill = PatternFill("solid", fgColor="FFF2CC")
        for k in range(1, 5):
            wi.cell(i, k).border = THIN
    for k, wdt in enumerate((46, 12, 14, 46), 1):
        wi.column_dimensions[get_column_letter(k)].width = wdt
    wi.freeze_panes = "A2"

    # ---- 信号交叉表 / 逐月汇总(公式)----
    wx = wb.create_sheet("信号交叉表")
    wx.append(["信号类型 \\ 平台信号", "平台突破（研究）", "平台观察", "无平台信号", "合计"])
    for k in range(1, 6):
        wx.cell(1, k).fill = NAVY
        wx.cell(1, k).font = HF
        wx.column_dimensions[get_column_letter(k)].width = 18
    for i, v in enumerate(("强确认", "标准确认", "观察级", "无信号"), 2):
        wx.cell(i, 1, v)
        for k, pv in enumerate(("平台突破（研究）", "平台观察", "无平台信号"), 2):
            wx.cell(i, k, f'=COUNTIFS({rr(g)},"{v}",{rr(t_)},"{pv}")')
        wx.cell(i, 5, f"=SUM(B{i}:D{i})")
        for k in range(1, 6):
            wx.cell(i, k).border = THIN
    wx.cell(6, 1, "合计")
    for k in range(2, 6):
        wx.cell(6, k, f"=SUM({get_column_letter(k)}2:{get_column_letter(k)}5)")
    wx.cell(8, 1, "经验:强确认(RPS60≥90)与任何平台信号的交集通常为 0 —— "
                  "强确认要求正在强势拉升,平台状态要求已回踩 20 周线并缩量走平,"
                  "两者在构造上互斥。若这一格不为 0,先怀疑平台或分档的实现。")
    wx.cell(8, 1).font = Font(bold=True, color="C00000")

    wm = wb.create_sheet("逐月汇总")
    wm.append(["观察月(YYYY-MM)", "强确认", "标准确认", "观察级",
               "平台突破（研究）", "平台观察", "统一信号=1"])
    for k in range(1, 8):
        wm.cell(1, k).fill = NAVY
        wm.cell(1, k).font = HF
        wm.column_dimensions[get_column_letter(k)].width = 18
    wm.cell(2, 1, "在 A 列填入年月(如 2022-01),右侧公式自动统计")
    wm.cell(2, 1).font = Font(italic=True, color="808080")
    for i in range(3, 27):
        for k, v in ((2, "强确认"), (3, "标准确认"), (4, "观察级")):
            wm.cell(i, k, f'=IF($A{i}="","",'
                          f'COUNTIFS({rr(g)},"{v}",{rr(od)},">="&DATEVALUE($A{i}&"-01"),'
                          f'{rr(od)},"<"&EDATE(DATEVALUE($A{i}&"-01"),1)))')
        for k, v in ((5, "平台突破（研究）"), (6, "平台观察")):
            wm.cell(i, k, f'=IF($A{i}="","",'
                          f'COUNTIFS({rr(t_)},"{v}",{rr(od)},">="&DATEVALUE($A{i}&"-01"),'
                          f'{rr(od)},"<"&EDATE(DATEVALUE($A{i}&"-01"),1)))')
        wm.cell(i, 7, f'=IF($A{i}="","",B{i}+C{i})')
    f = f"{OUT}/signal_template_blank_v1.0.xlsx"
    wb.save(f)
    print(f"已生成空白模板 {f}")
    print(f"  工作表 {len(wb.sheetnames)} 张:{wb.sheetnames}")
    print(f"  字段 {len(DICT)} 项;检查 {len(checks)} 条;恒等式自检 {len(ids)} 条")


if __name__ == "__main__":
    main()
