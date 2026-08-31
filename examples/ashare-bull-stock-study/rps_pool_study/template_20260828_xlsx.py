"""§164 交付:2026-08-28 当日清单 + 锚点,整合成带分组表头与配色的 Excel。

**本脚本不做任何计算,只排版** —— 读 `template_20260828.csv` 与
`template_20260828_anchor.csv`,输出 `template_20260828.xlsx`。

配色对齐 Codex 工作簿的观感(深藏青表头 + 分组色带 + 分档底色),
**一律用静态样式,不用条件格式** —— 第一四九节在 WPS 里踩过白字白底看不见的坑。
"""

from __future__ import annotations

import os

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.environ.get("OXQ_OUT_DIR", "/home/user/oxq-panel")
NAVY = PatternFill("solid", fgColor="17365D")
HF = Font(color="FFFFFF", bold=True, size=10)
BAND = {"基础与样本": "595959", "信号输出": "1F6FC5", "平台与质量": "1F7A5C",
        "价格与趋势": "8A6D1F"}
GRP = [("基础与样本", ["样本类型", "观察日期", "股票代码", "股票名称", "收盘价"]),
       ("信号输出", ["统一信号", "信号类型", "信号理由", "首次触发日期",
                  "连续确认天数", "触发状态"]),
       ("平台与质量", ["平台信号", "周线多头排列", "案例展示分层_质量",
                   "案例辅助标签_周线五态", "平台深度", "平台缩量比", "平台收敛比",
                   "R09核心质量分"]),
       ("价格与趋势", ["距一年低点涨幅", "近120日收益", "MA20持续度", "RPS50",
                   "RPS60", "RPS250", "距一年高点价格差"])]
COLS = [c for _, cs in GRP for c in cs]
FILL = {
    "信号类型": {"强确认": "FFC7CE", "标准确认": "FFEB9C", "观察级": "DDEBF7"},
    "平台信号": {"平台突破（研究）": "F8CBAD", "平台观察": "FFF2CC"},
    "案例展示分层_质量": {"高": "C6EFCE", "低": "FFC7CE", "缺失": "E7E6E6"},
    "案例辅助标签_周线五态": {"多头趋势": "C6EFCE", "突破启动": "F8CBAD",
                     "回踩修复": "FFF2CC", "均线蓄势": "DDEBF7",
                     "弱势结构": "FFC7CE", "未知": "E7E6E6"},
    "触发状态": {"新触发": "C6EFCE", "持续": "DDEBF7"},
}
PCT = ("距一年低点涨幅", "近120日收益", "MA20持续度", "距一年高点价格差",
       "平台深度")
THIN = Border(*[Side("thin", color="D9D9D9")] * 4)


def sheet(wb, name, df, note=None):
    ws = wb.create_sheet(name[:31])
    r0 = 1
    if note:
        ws.cell(1, 1, note).font = Font(bold=True, color="17365D", size=10)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        r0 = 2
    i = 1
    for g, cs in GRP:
        cs2 = [c for c in cs if c in df.columns]
        if not cs2:
            continue
        ws.merge_cells(start_row=r0, start_column=i, end_row=r0,
                       end_column=i + len(cs2) - 1)
        c = ws.cell(r0, i, g)
        c.fill = PatternFill("solid", fgColor=BAND[g])
        c.font = HF
        c.alignment = Alignment("center", "center")
        i += len(cs2)
    for k, col in enumerate(df.columns, 1):
        c = ws.cell(r0 + 1, k, col)
        c.fill = NAVY
        c.font = HF
        c.alignment = Alignment("center", "center", wrap_text=True)
    for _, r in df.iterrows():
        ws.append([None if (isinstance(v, float) and not np.isfinite(v)) else v
                   for v in r.tolist()])
    n = len(df)
    for k, col in enumerate(df.columns, 1):
        wmax = df[col].astype(str).str.len().max()
        wmax = 8 if not np.isfinite(wmax) else int(wmax)
        ws.column_dimensions[get_column_letter(k)].width = max(
            9, min(17, wmax + 2))
        if col in PCT:
            for rr in range(r0 + 2, r0 + 2 + n):
                ws.cell(rr, k).number_format = "0.0%"
        if col in ("RPS50", "RPS60", "RPS250", "R09核心质量分", "平台缩量比", "平台收敛比"):
            for rr in range(r0 + 2, r0 + 2 + n):
                ws.cell(rr, k).number_format = "0.00"
    idxmap = {c: k for k, c in enumerate(df.columns, 1)}
    for rr in range(r0 + 2, r0 + 2 + n):
        for k in range(1, len(df.columns) + 1):
            ws.cell(rr, k).border = THIN
        for col, mp in FILL.items():
            k = idxmap.get(col)
            if not k:
                continue
            v = ws.cell(rr, k).value
            f = mp.get(str(v))
            if f:
                ws.cell(rr, k).fill = PatternFill("solid", fgColor=f)
                if col == "触发状态" and v == "新触发":
                    ws.cell(rr, k).font = Font(bold=True)
    ws.freeze_panes = ws.cell(r0 + 2, 5)
    ws.auto_filter.ref = (f"A{r0+1}:{get_column_letter(len(df.columns))}"
                          f"{r0+1+n}")
    return ws


def _codex_agree():
    try:
        v = pd.read_csv(f"{OUT}/template_20260828_vs_codex.csv")
        v = v[v["他"].isin(["强确认", "标准确认", "观察级"])]
        return int((v["他"] == v["我"]).sum())
    except OSError:
        return -1


def main():
    a = pd.read_csv(f"{OUT}/template_20260828.csv", dtype={"股票代码": str})
    a["股票代码"] = a["股票代码"].str.zfill(6)
    a = a[COLS]
    k = pd.read_csv(f"{OUT}/template_20260828_anchor.csv", dtype={"代码": str})
    k["代码"] = k["代码"].str.zfill(6)
    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    # 强确认 × 平台信号 恒为 0(RPS60≥90 与"已回踩20周线并缩量走平"构造上互斥),
    # 故重点档取「统一信号=1(强确认或标准确认)且 平台突破」。
    core = a[(a["统一信号"] == 1) & (a["平台信号"] == "平台突破（研究）")]
    xt = pd.crosstab(a["信号类型"], a["平台信号"]).reset_index()
    lines = [
        ("2026-08-28 当日带信号清单 —— 按 Codex 模板口径,本地全市场面板", 0),
        (f"共 {len(a):,} 行、{a['股票代码'].nunique():,} 只｜"
         f"观察日 2026-08-28(单日全市场)｜面板 3,316 交易日 × 5,232 只｜行情与财务已更新至 2026-08-28", 1),
        ("", 0),
        ("■ 口径已与 Codex 对齐(2026-08-28 回函)", 0),
        ("信号类型(与 Codex RPS50 版 662 池逐只)86/86 = 100.0%  ✓", 1),
        ("平台信号(v0.4 案例)275/276 = 99.6%  ✓", 1),
        ("周线五态(v0.4 案例)276/276 = 100.0%  ✓", 1),
        ("**分档已统一改用 RPS50**(与 Codex rule_version=claude_rps50_weekly_v1.0 对齐);"
         "RPS60 仍照常输出供交叉核对。", 1),
        ("**平台强势日仍用 RPS60** —— 宇通 42天/2023-10-17 与第一五五节全部验证都建立在"
         "RPS60 上,换掉即作废,故未一并更改,待确认。", 1),
        ("R09 核心质量分:回函已确认案例表那一列未严格复用正式 R09,"
         "本表用的是正式 eligibility + 1%/99% 缩尾 + 四项等权,不与案例对齐。", 1),
        ("", 0),
        ("■ 颜色怎么看", 0),
        ("信号类型:强确认=粉  标准确认=黄  观察级=蓝  无信号=白", 1),
        ("平台信号:平台突破(研究)=橙  平台观察=浅黄  无平台信号=白", 1),
        ("质量分层:高=绿  中=白  低=粉  缺失=灰", 1),
        ("周线五态:多头趋势=绿  突破启动=橙  回踩修复=黄  均线蓄势=蓝  弱势结构=粉", 1),
        ("触发状态:新触发=绿加粗  持续=蓝  未触发=白", 1),
        ("", 0),
        ("■ 工作表", 0),
        (f"重点·双信号     {len(core):,} 行 —— 统一信号=1 且 平台突破", 1),
        ("信号交叉表      信号类型 × 平台信号 的行数交叉 —— "
         "强确认 × 任何平台信号恒为 0", 1),
        (f"强确认          {int((a['信号类型']=='强确认').sum()):,} 行", 1),
        (f"平台突破        {int((a['平台信号']=='平台突破（研究）').sum()):,} 行", 1),
        (f"全部清单        {len(a):,} 行(月末观察 + 平台突破日)", 1),
        ("逐月汇总        每月各档只数", 1),
        ("锚点对照        272 行案例逐行比对,一致=绿、不一致=粉", 1),
        ("检查            数量勾稽", 1),
        ("", 0),
        ("■ 边界(必须读)", 0),
        ("平台信号仍为 WATCHLIST 研究状态,不是买点。Codex 回函已指出胜宏科技 "
         "2023-09-12 是典型假突破,平台突破尚缺 RPS/K线质量/次日确认等过滤。", 1),
        ("「案例展示分层_质量」「案例辅助标签_周线五态」按回函要求降级命名,"
         "均非正式规则;正式的周线字段是二元「周线多头排列」。", 1),
        ("X01 的「短历史 listed_days ≥ 121」一档,回函未说明距一年低点(250日)"
         "在短历史下如何计算,本表一律用 ≥250。", 1),
        ("", 0),
        ("注意:这是候选筛选与状态标记,不是买入指令,不构成投资建议。", 0)]
    for t, i in lines:
        ws.append([("    " * i) + t])
    ws["A1"].fill = NAVY
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    for r in (4, 10, 17, 26):
        ws.cell(r, 1).font = Font(bold=True, color="17365D")
    ws.cell(len(lines), 1).fill = PatternFill("solid", fgColor="FFC7CE")
    ws.column_dimensions["A"].width = 108

    sheet(wb, f"重点·双信号{len(core)}", core,
          "统一信号=1(强确认或标准确认)且 平台突破(研究)"
          "—— 注:强确认 × 任何平台信号恒为 0,两者在构造上互斥")
    wsx = wb.create_sheet("信号交叉表")
    wsx.append(list(xt.columns))
    for c in range(1, len(xt.columns) + 1):
        wsx.cell(1, c).fill = NAVY
        wsx.cell(1, c).font = HF
        wsx.column_dimensions[get_column_letter(c)].width = 16
    for _, r in xt.iterrows():
        wsx.append(r.tolist())
    wsx.append([])
    wsx.append(["强确认(RPS60≥90)与任何平台信号的交集恒为 0:"
                "强确认要求股票正在强势拉升,平台状态要求它已回踩 20 周线并缩量走平,"
                "两者在构造上互斥。"])
    wsx.cell(wsx.max_row, 1).font = Font(bold=True, color="C00000")
    sheet(wb, "强确认", a[a["信号类型"] == "强确认"])
    sheet(wb, "平台突破", a[a["平台信号"] == "平台突破（研究）"])
    sheet(wb, "全部清单", a)
    m = a.groupby(["观察日期", "信号类型"]).size().unstack(fill_value=0)
    p = a.groupby(["观察日期", "平台信号"]).size().unstack(fill_value=0)
    mm = m.join(p, how="outer", rsuffix="_平台").fillna(0).astype(int).reset_index()
    ws2 = wb.create_sheet("逐月汇总")
    ws2.append(list(mm.columns))
    for c in range(1, len(mm.columns) + 1):
        ws2.cell(1, c).fill = NAVY
        ws2.cell(1, c).font = HF
        ws2.cell(1, c).alignment = Alignment("center", "center", wrap_text=True)
        ws2.column_dimensions[get_column_letter(c)].width = 14
    for _, r in mm.iterrows():
        ws2.append(r.tolist())
    ws2.freeze_panes = "B2"

    ws3 = wb.create_sheet("锚点对照")
    ws3.append(list(k.columns))
    for c in range(1, len(k.columns) + 1):
        ws3.cell(1, c).fill = NAVY
        ws3.cell(1, c).font = HF
        ws3.column_dimensions[get_column_letter(c)].width = 14
    for _, r in k.iterrows():
        ws3.append([None if (isinstance(v, float) and not np.isfinite(v)) else v
                    for v in r.tolist()])
    cm = {c: i for i, c in enumerate(k.columns, 1)}
    for rr in range(2, len(k) + 2):
        for f in ("信号类型", "平台信号", "周线"):
            if f"他_{f}" not in cm or f"我_{f}" not in cm:
                continue
            i1, i2 = cm.get(f"他_{f}"), cm.get(f"我_{f}")
            if not (i1 and i2):
                continue
            same = str(ws3.cell(rr, i1).value) == str(ws3.cell(rr, i2).value)
            for i in (i1, i2):
                ws3.cell(rr, i).fill = PatternFill(
                    "solid", fgColor="C6EFCE" if same else "FFC7CE")
    ws3.freeze_panes = "C2"

    try:
        vc = pd.read_csv(f"{OUT}/template_20260828_vs_codex.csv",
                         dtype={"代码": str})
        vc["代码"] = vc["代码"].str.zfill(6)
        ws5 = wb.create_sheet("与Codex逐只")
        ws5.append(list(vc.columns))
        for c in range(1, len(vc.columns) + 1):
            ws5.cell(1, c).fill = NAVY
            ws5.cell(1, c).font = HF
            ws5.column_dimensions[get_column_letter(c)].width = 14
        for _, r in vc.iterrows():
            ws5.append([None if (isinstance(v, float) and not np.isfinite(v)) else v
                        for v in r.tolist()])
        ci = {c: i for i, c in enumerate(vc.columns, 1)}
        for rr in range(2, len(vc) + 2):
            a_, b_ = ws5.cell(rr, ci["他"]).value, ws5.cell(rr, ci["我"]).value
            if a_ in ("强确认", "标准确认", "观察级"):
                f = "C6EFCE" if str(a_) == str(b_) else "FFC7CE"
                for i in (ci["他"], ci["我"]):
                    ws5.cell(rr, i).fill = PatternFill("solid", fgColor=f)
        ws5.freeze_panes = "B2"
    except OSError:
        pass

    chk = pd.DataFrame([
        {"检查项": "总行数", "值": len(a)},
        {"检查项": "涉及股票只数", "值": a["股票代码"].nunique()},
        {"检查项": "当日观察行", "值": int((a["样本类型"] == "当日观察").sum())},
        {"检查项": "强确认", "值": int((a["信号类型"] == "强确认").sum())},
        {"检查项": "标准确认", "值": int((a["信号类型"] == "标准确认").sum())},
        {"检查项": "观察级", "值": int((a["信号类型"] == "观察级").sum())},
        {"检查项": "统一信号=1(=强确认+标准确认)", "值": int(a["统一信号"].sum())},
        {"检查项": "  勾稽:强确认+标准确认",
         "值": int((a["信号类型"].isin(["强确认", "标准确认"])).sum())},
        {"检查项": "平台突破（研究）",
         "值": int((a["平台信号"] == "平台突破（研究）").sum())},
        {"检查项": "平台观察", "值": int((a["平台信号"] == "平台观察").sum())},
        {"检查项": "重点·双信号(统一信号=1 且 平台突破)", "值": len(core)},
        {"检查项": "  强确认 × 平台突破(构造上互斥,应为 0)",
         "值": int(((a["信号类型"] == "强确认")
                   & (a["平台信号"] == "平台突破（研究）")).sum())},
        {"检查项": "周线多头排列为真", "值": int(a["周线多头排列"].sum())},
        {"检查项": "锚点行数", "值": len(k)},
        {"检查项": "锚点二·与Codex信号类型一致(其有信号 86 只)",
         "值": _codex_agree()},
        {"检查项": "锚点·平台信号一致",
         "值": int((k["他_平台信号"] == k["我_平台信号"]).sum())},
        {"检查项": "锚点·周线五态一致",
         "值": int((k["他_周线"] == k["我_周线"]).sum())}])
    ws4 = wb.create_sheet("检查")
    ws4.append(["检查项", "值"])
    for c in (1, 2):
        ws4.cell(1, c).fill = NAVY
        ws4.cell(1, c).font = HF
    for _, r in chk.iterrows():
        ws4.append(r.tolist())
    ws4.column_dimensions["A"].width = 34
    ws4.column_dimensions["B"].width = 12
    f = f"{OUT}/template_20260828.xlsx"
    wb.save(f)
    print(f"已生成 {f}")
    print(f"  工作表 {len(wb.sheetnames)} 张:{wb.sheetnames}")
    print(f"  重点·双信号 {len(core)} 行;强确认 "
          f"{int((a['信号类型']=='强确认').sum()):,};全部 {len(a):,}")


if __name__ == "__main__":
    main()
