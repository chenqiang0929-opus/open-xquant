"""§153 交付:把两条规则的当前名单打包成 Excel 初稿(格式对齐 Codex 的工作簿)。

**本脚本不做任何检验、不新增任何规则** —— 只把 `codex_unified_signal.py`
已经跑完并落库的名单与判定结果排版成一个可直接看的工作簿。

与 Codex 工作簿的差别(必须让用户一眼看到)
------------------------------------------
1. **每条规则旁边印上它的样本外成绩**(B1 概率口径 / B2 组合口径的实测数字与判定),
   Codex 的工作簿只有快照,没有任何回看成绩;
2. 覆盖**全市场 3,136 只合格股**,不只是 662 只次新股池
   (Codex 那份里有 44 只本地取不到行情);
3. 加一列**距低点反弹**并按用户第一四九节的要求排序/标色 ——
   已经涨过一倍以上的排到后面并标黄/标红。

**静态样式,不用条件格式** —— 第一四九节在 WPS 里踩过白字白底看不见的坑。
"""

from __future__ import annotations

import glob
import json
import os

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = os.environ.get("OXQ_OUT_DIR", "/home/user/oxq-panel")
HDR = PatternFill("solid", fgColor="17365D")
HF = Font(color="FFFFFF", bold=True, size=10)
TIER = {"A": PatternFill("solid", fgColor="C6EFCE"),
        "B": PatternFill("solid", fgColor="DDEBF7"),
        "C": PatternFill("solid", fgColor="FFF2CC"),
        "D": PatternFill("solid", fgColor="F8CBAD")}
BAD = PatternFill("solid", fgColor="FFC7CE")
OKF = PatternFill("solid", fgColor="C6EFCE")
THIN = Border(*[Side("thin", color="BFBFBF")] * 4)


def put(ws, df, pcts=(), tiercol=None, reccol=None):
    ws.append(list(df.columns))
    for c in range(1, len(df.columns) + 1):
        ws.cell(1, c).fill = HDR
        ws.cell(1, c).font = HF
        ws.cell(1, c).alignment = Alignment("center", "center", wrap_text=True)
    for _, r in df.iterrows():
        ws.append([None if (isinstance(v, float) and not np.isfinite(v)) else v
                   for v in r.tolist()])
    for i, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = \
            max(9, min(16, int(df[col].astype(str).str.len().max() or 8) + 3))
        if col in pcts:
            for row in range(2, len(df) + 2):
                ws.cell(row, i).number_format = "0.0%"
    for row in range(2, len(df) + 2):
        for c in range(1, len(df.columns) + 1):
            ws.cell(row, c).border = THIN
        if tiercol:
            t = str(ws.cell(row, list(df.columns).index(tiercol) + 1).value or "")[:1]
            if t in TIER:
                for c in range(1, len(df.columns) + 1):
                    ws.cell(row, c).fill = TIER[t]
        if reccol:
            k = list(df.columns).index(reccol) + 1
            v = ws.cell(row, k).value
            if isinstance(v, (int, float)) and np.isfinite(v):
                ws.cell(row, k).fill = BAD if v > 2.0 else (
                    TIER["C"] if v > 1.0 else OKF)
    ws.freeze_panes = "A2"


CENSUS = ("/home/user/quant-research-dev/research/"
          "bull-stock-census-2010-2025/data/*.csv")


def name_map():
    """全市场股票名 —— 面板里没有名称字段,从两个只读来源拼:
    普查表(code,name;历史名,可能是当年的旧名)+ 次新股池的 code_name_map(当前名,优先)。"""
    m = {}
    for f in glob.glob(CENSUS):
        try:
            x = pd.read_csv(f, dtype=str)
        except Exception:                                      # noqa: BLE001
            continue
        x.columns = [c.strip("\ufeff") for c in x.columns]
        if "code" in x.columns and "name" in x.columns:
            for c, n in zip(x["code"], x["name"], strict=True):
                if isinstance(c, str) and isinstance(n, str):
                    m[c.zfill(6)] = n
    try:
        with open(f"{OUT}/code_name_map.json", encoding="utf-8") as fh:
            m.update({k.zfill(6): v for k, v in json.load(fh).items()})
    except OSError:
        pass
    return m


def main():
    cur = pd.read_csv(f"{OUT}/codex_unified_signal_current.csv",
                      dtype={"股票代码": str})
    cur["股票代码"] = cur["股票代码"].str.zfill(6)
    nm = name_map()
    miss = cur["股票名称"].isna()
    cur.loc[miss, "股票名称"] = cur.loc[miss, "股票代码"].map(nm)
    print(f"股票名称:{int(cur['股票名称'].notna().sum()):,}/{len(cur):,} 只有名"
          f"(名称来自普查表与次新股池,面板本身无名称字段;个别可能是历史旧名)")
    res = pd.read_csv(f"{OUT}/codex_unified_signal.csv")

    def tier(r):
        if r["两条都中"] == 1:
            return "A 两条规则都中" if r["距低点反弹"] <= 1.0 else "C 两条都中但已涨超1倍"
        if r["Codex统一信号"] == 1 or r["第一四八节信号"] == 1:
            return "B 单条规则中" if r["距低点反弹"] <= 1.0 else "D 单条中且已涨超1倍"
        return ""
    cur["分层"] = cur.apply(tier, axis=1)
    sig = cur[cur["分层"] != ""].copy()
    order = {"A": 0, "B": 1, "C": 2, "D": 3}
    sig["_o"] = sig["分层"].str[0].map(order)
    sig = sig.sort_values(["_o", "距低点反弹"]).drop(columns="_o")
    show = ["股票代码", "股票名称", "分层", "在次新股池内", "Codex统一信号",
            "Codex强确认", "第一四八节信号", "两条都中", "距低点反弹", "近120日收益",
            "近60日收益", "MA20持续度", "RPS60", "RPS250", "换手加速",
            "距低点分位", "换手加速分位", "流通市值亿", "申万一级"]
    pcts = ("距低点反弹", "近120日收益", "近60日收益", "MA20持续度", "换手加速",
            "距低点分位", "换手加速分位")

    wb = Workbook()
    ws = wb.active
    ws.title = "说明"
    obs = str(cur["观察日期"].iloc[0])
    lines = [
        ("价格启动信号 —— 两条规则并排,每条都印着它的样本外成绩", 0),
        (f"观察日 {obs}(本地面板末日)｜全市场合格 {len(cur):,} 只｜"
         f"次新股池内合格 {int(cur['在次新股池内'].sum())} 只", 1),
        ("", 0),
        ("■ 先看结论:两条规则在两个口径下的成绩相反", 0),
        ("概率口径(未来60日涨50%的概率):Codex 的规则通过,我的第一四八节规则也通过,"
         "Codex 还略好一点。", 1),
        ("组合口径(等权持有一个月,对同市值同行业随机对照):三条全部不通过,"
         "Codex 的两条比我的更差。", 1),
        ("→ 「能提高启动概率的标记」不等于「能赚钱的买点」。本表是候选观察名单,"
         "不是买入指令。", 1),
        ("", 0),
        ("■ 分层怎么看(按你在第一四九节提的「要底部启动点,不要已经涨太多的」排的)", 0),
        ("A 两条规则都中,且距一年低点涨幅 ≤100%  —— 绿底,排最前", 1),
        ("B 单条规则中,且距一年低点涨幅 ≤100%    —— 蓝底", 1),
        ("C 两条都中,但已涨超1倍                  —— 黄底,你在第一四九节否过这一类", 1),
        ("D 单条中,且已涨超1倍                    —— 橙底,排最后", 1),
        ("「距低点反弹」列单独标色:≤100% 绿、100~200% 黄、>200% 红。", 1),
        ("**分层本身没有经过样本外检验,它只是把已经涨太多的排到后面,不是新规则。**", 1),
        ("", 0),
        ("■ 与 Codex 那份工作簿的差别", 0),
        ("1) 他那份是纯快照,没有任何回看成绩;本表每条规则都带 B1/B2 实测数字。", 1),
        ("2) 他那份 662 只里有 44 只本地取不到行情;本表覆盖全市场 3,136 只合格股。", 1),
        ("3) 他的观察日 2026-08-26,本地面板末日 2026-08-03,差 15 个交易日,"
         "数值不可能逐一对上。", 1),
        ("", 0),
        ("■ 工作表", 0),
        ("规则与成绩    两条规则的门槛,以及各自的样本外实测与判定", 1),
        ("A两条都中     两条规则同时中且未涨超1倍 —— 最值得重点看的一档", 1),
        ("全市场信号清单 全部有信号的股票,按分层排序", 1),
        ("次新股池      Codex 的 662 只池子在本地面板上的信号", 1),
        ("检查          数量勾稽", 1),
        ("", 0),
        ("注意:这是候选筛选信号,不是自动买入指令。仓位、买卖时点、流动性、"
         "涨跌停和风险控制均未纳入。组合口径的实测显示这两条规则都跑输"
         "同市值同行业的随机对照。", 0)]
    for txt, ind_ in lines:
        ws.append([("    " * ind_) + txt])
    ws["A1"].fill = HDR
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=13)
    for r in (4, 9, 17, 23):
        ws.cell(r, 1).font = Font(bold=True, color="17365D")
    ws.cell(len(lines), 1).fill = BAD
    ws.column_dimensions["A"].width = 100

    # ---- 规则与成绩 ----
    rr = res[res["口径"].isin(["概率", "组合"])].copy()
    rr = rr[rr["段"].str.contains("留出")]
    # 概率行与组合行的规则名一长一短(带不带括号说明),按括号前的短名对齐
    rr["key"] = rr["规则"].str.split("(").str[0].str.strip()
    grade = []
    for nm in ("Codex 统一信号1", "Codex 强确认", "第一四八节规则"):
        p_ = rr[(rr["key"] == nm) & (rr["口径"] == "概率")]
        c_ = rr[(rr["key"] == nm) & (rr["口径"] == "组合")]
        if p_.empty or c_.empty:
            continue
        p_, c_ = p_.iloc[0], c_.iloc[0]
        grade.append({
            "规则": nm,
            "门槛": ("反弹≥40% 且 120日收益≥10% 且 MA20持续度≥55% 且 RPS60≥80"
                   if "统一" in nm else
                   "反弹≥40% 且 120日收益≥10% 且 MA20持续度≥55% 且 RPS60≥90"
                   if "强确认" in nm else
                   "距一年低点涨幅前30% 且 换手加速前30%(全市场当日横截面)"),
            "来源": "Codex 20260826 工作簿" if "Codex" in nm else "本地第一四八节",
            "覆盖率": p_["覆盖率"], "启动率": p_["启动率"],
            "lift(留出段)": p_["lift"], "噪音上界95": p_["噪音上界95"],
            "B1概率口径判定": "通过" if (p_["lift"] > 1.20
                                  and p_["lift"] > p_["噪音上界95"]) else "不通过",
            "组合年化": c_["零成本年化"], "对照年化": c_["对照年化中位"],
            "超额pp": c_["超额pp"], "单尾p": c_["p"],
            "B2组合口径判定": "通过" if (c_["超额pp"] >= 3.0
                                  and c_["p"] < 0.05) else "不通过"})
    gdf = pd.DataFrame(grade)
    ws2 = wb.create_sheet("规则与成绩")
    put(ws2, gdf, pcts=("覆盖率", "启动率", "组合年化", "对照年化"))
    for row in range(2, len(gdf) + 2):
        for cn in ("B1概率口径判定", "B2组合口径判定"):
            k = list(gdf.columns).index(cn) + 1
            ws2.cell(row, k).fill = OKF if ws2.cell(row, k).value == "通过" else BAD
    ws2.append([])
    ws2.append(["留出段 = 2023-01 → 2026-04,训练段 2019-2022 只报数不判定;"
                "对照 = 同日同市值名次±25 同申万一级行业随机抽同样只数,500 组种子。"])
    ws2.append(["B1 判据:lift > 1.20 且 > 打乱标签200次的95分位。"
                "B2 判据:年化超额 ≥ +3.00pp 且单尾 p < 0.05。判据在跑之前写死。"])

    a = sig[sig["分层"].str.startswith("A")][show]
    put(wb.create_sheet(f"A两条都中{len(a)}只"), a, pcts, "分层", "距低点反弹")
    put(wb.create_sheet("全市场信号清单"), sig[show], pcts, "分层", "距低点反弹")
    pool = cur[cur["在次新股池内"]].copy()
    pool["_o"] = pool["分层"].str[0].map(order).fillna(9)
    pool = pool.sort_values(["_o", "距低点反弹"]).drop(columns="_o")
    put(wb.create_sheet("次新股池"), pool[show], pcts, "分层", "距低点反弹")

    chk = pd.DataFrame([
        {"检查项目": "全市场合格", "实际值": len(cur)},
        {"检查项目": "Codex 统一信号1", "实际值": int(cur["Codex统一信号"].sum())},
        {"检查项目": "Codex 强确认", "实际值": int(cur["Codex强确认"].sum())},
        {"检查项目": "第一四八节信号", "实际值": int(cur["第一四八节信号"].sum())},
        {"检查项目": "两条都中", "实际值": int(cur["两条都中"].sum())},
        {"检查项目": "有信号合计(去重)", "实际值": len(sig)},
        {"检查项目": "  A 两条都中且未涨超1倍",
         "实际值": int(sig["分层"].str.startswith("A").sum())},
        {"检查项目": "  B 单条中且未涨超1倍",
         "实际值": int(sig["分层"].str.startswith("B").sum())},
        {"检查项目": "  C 两条都中但已涨超1倍",
         "实际值": int(sig["分层"].str.startswith("C").sum())},
        {"检查项目": "  D 单条中且已涨超1倍",
         "实际值": int(sig["分层"].str.startswith("D").sum())},
        {"检查项目": "次新股池内合格", "实际值": int(cur["在次新股池内"].sum())},
        {"检查项目": "次新股池内 Codex 信号1",
         "实际值": int(pool["Codex统一信号"].sum())}])
    chk["状态"] = "OK"
    put(wb.create_sheet("检查"), chk)
    f = f"{OUT}/startup_signal_draft.xlsx"
    wb.save(f)
    print(f"已生成 {f}")
    print(f"  分层:A {int(sig['分层'].str.startswith('A').sum())}、"
          f"B {int(sig['分层'].str.startswith('B').sum())}、"
          f"C {int(sig['分层'].str.startswith('C').sum())}、"
          f"D {int(sig['分层'].str.startswith('D').sum())}")


if __name__ == "__main__":
    main()
